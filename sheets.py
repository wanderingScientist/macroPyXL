from PyQt6.QtWidgets import (QApplication, QMainWindow, QTableWidget,
                             QTableWidgetItem, QVBoxLayout, QWidget, QTextEdit,
                             QPushButton, QMessageBox, QInputDialog, QSplitter,
                             QHBoxLayout, QFileDialog, QColorDialog)
from PyQt6.QtGui import QColor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
import sys
import json
import numpy as np
from typing import NamedTuple
import subprocess


class internalSlicer:

    def __getitem__(self, key):
        return key


class internalSliceRange:

    def __init__(self):
        self.slicer = internalSlicer()

    def __getitem__(self, key):
        s = self.slicer[key]
        if isinstance(s, slice):
            return range(s.start, s.stop, s.step or 1)
        else:
            return s


class baseUtils(NamedTuple):
    S: internalSlicer = internalSlicer()
    R: internalSliceRange = internalSliceRange()


BU = baseUtils()


class SpreadsheetApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("PyQt6 Spreadsheet with Macros")
        self.setGeometry(100, 100, 800, 600)

        self.cell_globals = {}
        self.cell_colors = {}  # Store cell colors separately to prevent reset
        self.text_colors = {}  # Store cell colors separately to prevent reset
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.splitter = QSplitter()

        self.table = QTableWidget(10, 10)
        self.table.cellChanged.connect(self.evaluate_cell)
        self.update_headers()
        self.initialize_cells()
        self.splitter.addWidget(self.table)

        self.macro_editor = QTextEdit()
        self.macro_editor.setPlaceholderText("Write your macros here...")
        self.splitter.addWidget(self.macro_editor)
        self.splitter.setSizes([500, 300
                                ])  # Initial sizes of table and macro editor

        layout.addWidget(self.splitter)

        button_layout = QHBoxLayout()

        self.run_button = QPushButton("Apply Macros")
        self.run_button.clicked.connect(self.apply_macros)
        button_layout.addWidget(self.run_button)

        self.save_macro_file_button = QPushButton("Save Macros")
        self.save_macro_file_button.clicked.connect(self.save_macros_to_file)
        button_layout.addWidget(self.save_macro_file_button)

        self.load_macro_file_button = QPushButton("Load Macros")
        self.load_macro_file_button.clicked.connect(self.load_macros_from_file)
        button_layout.addWidget(self.load_macro_file_button)

        self.resize_button = QPushButton("Resize")
        self.resize_button.clicked.connect(self.resize_table)
        button_layout.addWidget(self.resize_button)

        self.delete_selection_button = QPushButton("Delete Row/Column")
        self.delete_selection_button.clicked.connect(self.delete_row_or_column)
        button_layout.addWidget(self.delete_selection_button)

        self.clear_cells_button = QPushButton("Clear")
        self.clear_cells_button.clicked.connect(self.clear_selected_cells)
        button_layout.addWidget(self.clear_cells_button)

        self.change_color_button = QPushButton("Change Cell/Text Color")
        self.change_color_button.clicked.connect(self.change_cell_text_color)
        button_layout.addWidget(self.change_color_button)

        self.save_sheet_button = QPushButton("Save Sheet")
        self.save_sheet_button.clicked.connect(self.save_file)
        button_layout.addWidget(self.save_sheet_button)

        self.load_sheet_button = QPushButton("Load Sheet")
        self.load_sheet_button.clicked.connect(self.load_file)
        button_layout.addWidget(self.load_sheet_button)

        layout.addLayout(button_layout)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def show_error_dialog(self, message):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        msg_box.setText("An error occurred")
        msg_box.setInformativeText(message)
        msg_box.setWindowTitle("Error")
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def get_cell_value(self, row, col):
        item = self.table.item(row, col)
        return item.text() if item else ""

    def set_cell_color(self, row, col, r, g, b):
        self.cell_colors[(row, col)] = QColor(r, g, b)
        item = self.table.item(row, col)
        if not item:
            item = QTableWidgetItem("")
            self.table.setItem(row, col, item)
        item.setBackground(self.cell_colors[(row, col)])

    def set_cell_text_color(self, row, col, r, g, b):
        self.cell_colors[(row, col)] = QColor(r, g, b)
        item = self.table.item(row, col)
        if not item:
            item = QTableWidgetItem("")
            self.table.setItem(row, col, item)
        item.setForeground(self.cell_colors[(row, col)])

    def get_cell_values(self, rows, cols):
        rows = rows if isinstance(rows, range) else [rows]
        cols = cols if isinstance(cols, range) else [cols]
        values = [
            self.table.item(row, col).text() for row in rows for col in cols
        ]
        return values

    def _math_func(self, func, rows, cols):
        values = self.get_cell_values(rows, cols)
        values = list(map(float, values))
        return func(values)

    def sum(self, rows, cols):
        return self._math_func(np.sum, rows, cols)

    def mean(self, rows, cols):
        return self._math_func(np.mean, rows, cols)

    def std(self, rows, cols):
        return self._math_func(np.std, rows, cols)

    def apply_macros(self):
        macro_code = self.macro_editor.toPlainText()

        self.cell_globals[
            "get_cell_value"] = self.get_cell_value  # Expose function to macros
        self.cell_globals[
            "set_cell_color"] = self.set_cell_color  # Expose function to macros
        self.cell_globals["sum"] = self.sum
        self.cell_globals["mean"] = self.mean
        self.cell_globals["std"] = self.std
        self.cell_globals["np"] = np
        self.cell_globals["R"] = BU.R
        self.cell_globals["subprocess"] = subprocess

        try:
            exec(macro_code, self.cell_globals)
        except Exception as e:
            self.show_error_dialog(str(e))

    def evaluate_cell(self, row, col):
        item = self.table.item(row, col)
        if item:
            value = item.text()
            if value.startswith("="):
                try:
                    self.cell_globals["X"] = row  # Set dynamic row variable
                    self.cell_globals["Y"] = col  # Set dynamic column variable
                    expr = value[1:]
                    result = eval(expr, self.cell_globals)
                    self.table.blockSignals(True)
                    self.table.setItem(row, col, QTableWidgetItem(str(result)))
                    if (
                            row, col
                    ) in self.cell_colors:  # Restore color after evaluation
                        self.table.item(row, col).setBackground(
                            self.cell_colors[(row, col)])
                        self.table.item(row, col).setForeground(
                            self.text_colors[(row, col)])
                    self.table.blockSignals(False)
                except Exception as e:
                    self.table.blockSignals(True)
                    self.table.setItem(row, col, QTableWidgetItem("ERROR"))
                    self.table.blockSignals(False)
                    self.show_error_dialog(
                        f"Error evaluating cell ({row}, {col}): {e}")

    def initialize_cells(self):
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if not item:
                    item = QTableWidgetItem("")
                    self.table.setItem(row, col, item)
                item.setBackground(QColor(255, 255, 255))  # White background
                item.setForeground(QColor(0, 0, 0))  # Black text
                self.cell_colors[(row, col)] = QColor(255, 255, 255)
                self.text_colors[(row, col)] = QColor(0, 0, 0)

    def update_headers(self):
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        self.table.setHorizontalHeaderLabels(
            [str(i) for i in range(col_count)])
        self.table.setVerticalHeaderLabels([str(i) for i in range(row_count)])

    def resize_table(self):
        rows, ok1 = QInputDialog.getInt(self, "Resize Table",
                                        "Enter number of rows:",
                                        self.table.rowCount(), 1, 100)
        cols, ok2 = QInputDialog.getInt(self, "Resize Table",
                                        "Enter number of columns:",
                                        self.table.columnCount(), 1, 100)

        if ok1 and ok2:
            self.table.setRowCount(rows)
            self.table.setColumnCount(cols)
            self.update_headers()
            self.initialize_cells()

    def _as_dict(self):
        data = {}
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                value = item.text() if item else ""

                background_color = item.background().color().getRgb(
                )[:3] if item else (255, 255, 255)  # Default: White cell
                text_color = item.foreground().color().getRgb(
                )[:3] if item else (0, 0, 0)  # Default: Black text

                data[f"{row},{col}"] = {
                    "value": value,
                    "color": background_color,
                    "text_color": text_color
                }
        return data

    def save_file(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Save Spreadsheet", "",
            "JSON Files (*.json);;Excel Files (*.xlsx);;All Files (*)")

        if file_name:
            data = self._as_dict()
            if ".json" == file_name[-5:]:
                try:
                    with open(file_name, "w") as f:
                        json.dump(data, f, indent=4)
                except:
                    print(f"{data =}")
                    raise
            elif ".xlsx" == file_name[-5:]:
                save_json_to_xlsx(data, file_name)
            else:
                self.show_error_dialog(
                    f"Save format unsupported.\n{file_name}")

    def _load_json(self, data):

        self.table.blockSignals(True)
        self.cell_colors.clear()

        for key, cell_data in data.items():
            row, col = map(int, key.split(","))
            value = cell_data.get("value", "")
            color = cell_data.get("color", None)
            txtcolor = cell_data.get("text_color", None)

            _item = QTableWidgetItem(value)
            self.table.setItem(row, col, _item)

            item = self.table.item(row, col)

            if color:
                qcolor = QColor(*color)
                item.setBackground(qcolor)
                self.cell_colors[(row, col)] = qcolor
                self.set_cell_color(row, col, *cell_data["color"])

            if txtcolor:
                qcolor = QColor(*txtcolor)
                item.setForeground(qcolor)
                self.text_colors[(row, col)] = qcolor
                self.set_cell_text_color(row, col, *cell_data["text_color"])

        self.table.blockSignals(False)

    def load_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Load Spreadsheet", "",
            "JSON Files (*.json);;Excel Files (*.xlsx);;All Files (*)")

        if ".json" == file_name[-5:]:
            with open(file_name, "r") as f:
                data = json.load(f)
        elif ".xlsx" == file_name[-5:]:
            data = load_xlsx_to_dict(file_name)
        elif not file_name:
            return
        else:
            self.show_error_dialog(f"Load file {file_name} is unsupported")
            return
        self._load_json(data)

    def save_macros_to_file(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Save Macro File", "", "Python Files (*.py);;All Files (*)")
        if file_name:
            with open(file_name, "w") as f:
                f.write(self.macro_editor.toPlainText())

    def load_macros_from_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self, "Load Macro File", "", "Python Files (*.py);;All Files (*)")
        if file_name:
            with open(file_name, "r") as f:
                self.macro_editor.setPlainText(f.read())

    def clear_selected_cells(self):
        for item in self.table.selectedItems():
            item.setText("")
            item.setBackground(QColor(255, 255,
                                      255))  # Reset to white background
            item.setForeground(QColor(0, 0, 0))  # Reset to black text

    def change_cell_text_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            option, ok = QInputDialog.getItem(
                self, "Select Color Option", "Apply color to:",
                ["Cell Background", "Text Color"], 0, False)
            if ok:
                for item in self.table.selectedItems():
                    row, col = item.row(), item.column()
                    if option == "Cell Background":
                        item.setBackground(color)
                        self.cell_colors[(row, col)] = color
                    elif option == "Text Color":
                        item.setForeground(color)
                        self.text_colors[(row, col)] = color

    def delete_row_or_column(self):
        option, ok = QInputDialog.getItem(self, "Delete Row or Column",
                                          "Select option:", ["Row", "Column"],
                                          0, False)
        if ok:
            index, ok = QInputDialog.getInt(
                self, "Delete", f"Enter {option.lower()} index:", 0, 0,
                self.table.rowCount() -
                1 if option == "Row" else self.table.columnCount() - 1)
            if ok:
                if option == "Row":
                    self.table.removeRow(index)
                elif option == "Column":
                    self.table.removeColumn(index)
            self.update_headers()


def save_json_to_xlsx(json_data: dict, filename: str):
    wb = Workbook()
    ws = wb.active

    # Define a default border (thin lines)
    thin_border = Border(left=Side(style="thin", color="D3D3D3"),
                         right=Side(style="thin", color="D3D3D3"),
                         top=Side(style="thin", color="D3D3D3"),
                         bottom=Side(style="thin", color="D3D3D3"))

    for key, details in json_data.items():
        row, col = map(int, key.split(','))  # Convert "0,0" to row=0, col=0
        cell = ws.cell(row=row + 1,
                       column=col + 1)  # OpenPyXL uses 1-based indexing

        # Assign value
        cell.value = details["value"]

        # Assign background color
        r, g, b = details["color"]
        fill = PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                           end_color=f"{r:02X}{g:02X}{b:02X}",
                           fill_type="solid")
        cell.fill = fill

        # Assign text color
        tr, tg, tb = details["text_color"]
        font = Font(color=f"{tr:02X}{tg:02X}{tb:02X}")
        cell.font = font

        # Apply border
        cell.border = thin_border  # Ensures borders are visible in all programs

    # Save the Excel file
    wb.save(filename)


def rgb_from_hex(hex_color):
    """Convert a hex color (RRGGBB) to an RGB tuple."""
    hex_color = hex_color.lstrip("#")
    return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))


def load_xlsx_to_dict(filename):
    """Load an Excel file and return a dictionary with cell values, background colors, and text colors."""
    wb = load_workbook(filename)
    ws = wb.active
    data = {}

    for row in ws.iter_rows():
        for cell in row:
            row_idx, col_idx = cell.row - 1, cell.column - 1  # Convert to zero-based index
            key = f"{row_idx},{col_idx}"

            # Get cell value
            value = str(cell.value) if cell.value is not None else ""

            # Get background color (if set)
            bg_color = cell.fill.start_color.rgb  # Example: '00FFFF00'
            bg_rgb = rgb_from_hex(
                bg_color[2:]) if bg_color and bg_color != "00000000" else [
                    255, 255, 255
                ]  # Default to white

            # Get text color (if set)
            text_color = cell.font.color.rgb if cell.font and cell.font.color else "000000"
            text_rgb = rgb_from_hex(
                text_color[2:]) if text_color and text_color != "000000" else [
                    0, 0, 0
                ]  # Default to black

            # Store data in dictionary
            data[key] = {
                "value": value,
                "color": list(bg_rgb),
                "text_color": list(text_rgb)
            }

    return data


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SpreadsheetApp()
    window.show()
    sys.exit(app.exec())
